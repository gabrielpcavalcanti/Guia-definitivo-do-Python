import os
import openpyxl
from tkinter import filedialog

def choose_file():

    try:
        path_file = filedialog.askopenfilename()
        
        return path_file
    
    except OSError:
        print("Select a file")

spreadsheet_path = choose_file()
lista_dados = []

wb = openpyxl.load_workbook(spreadsheet_path)

ws = wb['Punditlink Data (2)']

wb.create_sheet("Organized Data", 1)

ws1 = wb['Organized Data']

wb.save(spreadsheet_path)


for row in ws.iter_rows(values_only=True):
    
    lista_dados_temporaria = list(row)
    lista_dados.append(lista_dados_temporaria)

lista_valores = list(range(16,len(lista_dados), 25))

ws1.append(lista_dados[14])

for row_lista in lista_dados:

    if lista_dados.index(row_lista) in lista_valores:
        ws1.append(row_lista)
        

wb.save(spreadsheet_path)
