import openpyxl

wb = openpyxl.load_workbook('cm.xlsx')

ws = wb['Sheet1']

for i in range(1,100):
    ws.cell(column=5, row=i).value = str(ws.cell(column=5, row=i).value)[:-3].replace(".",",")


wb.save('cm.xlsx')