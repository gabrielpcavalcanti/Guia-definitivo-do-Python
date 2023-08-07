import openpyxl

wb = openpyxl.load_workbook('e.xlsx')

ws = wb['Sheet1']

for i in range(1,100):

    
    try:
        ws.cell(column=1, row=i).value = str(ws.cell(column=1, row=i).value).replace(".",",")
        ws.cell(column=2, row=i).value = str(ws.cell(column=2, row=i).value).replace(".",",")

        

    
    except ValueError:
        continue
    except TypeError:
        continue

wb.save('novo_em_folha4.xlsx')